from openpyxl import load_workbook
import sys, os

class SapDataParser:
    #TODO calculate this
    upperRange = 0
    currentInfoArea = ''
    currentTable = ''
    outputRowNum = 2
    sourceFileName = ''
    outputFileName = 'output.xlsx'
    hasInfoArea = False
    hasTable = False 
    domain = 'SAFYR SAP Test'
    community = 'Technical Metadata Community'
    domainType = 'Physical Data Dictionary'
    def __init__(self,input, output):
        print('Loading Excel Files')
        self.sourceFileName = input
        self.wb = load_workbook(filename = input)
        self.ws = self.wb.active
        self.resetOutputFile(output)
        self.outputFileName = output
        self.output = load_workbook(output)
        self.sOutput = self.output.active
        self.buildFieldMap()
        self.buildHeaders()
    def resetOutputFile(self,fileName = 'output.xlsx'):
        pathName = '.\\emptyOutput\\' + fileName
        destination = '.\\'
        os.system('del {d}\{fn}'.format(d=destination,fn=fileName))
        os.system('copy {pn} {d}"'.format(pn=pathName,d=destination))
    def buildHeaders(self):    
        for k,v in self.fieldTemp.items():
            self.sOutput[v+'1'] = k
    def setDataLength(self, upperRange = 49646):
        #TODO: Calculate upperRange
        self.upperRange = upperRange + 1
    def buildFieldMap(self):
        print('Building Map')
        #src fields
        self.fieldSrc={}
        self.fieldSrc['Parent'] = 'A'
        self.fieldSrc['Child'] = 'B'
        self.fieldSrc['DD_TABLENAME'] = 'C'
        self.fieldSrc['LONG_DESC'] = 'D'
        self.fieldSrc['DD_TABLETYPE'] = 'E'
        self.fieldSrc['DD_FIELDNAME'] = 'F'
        self.fieldSrc['SHORT_DESC'] = 'G'
        self.fieldSrc['POSIT'] = 'H'
        self.fieldSrc['MANDATORY'] = 'I'
        self.fieldSrc['DD_DATATYPE_ERP'] = 'J'
        self.fieldSrc['DATA_LENGTH'] = 'K'
        self.fieldSrc['DATA_DECIMALS'] = 'L'
        self.fieldSrc['KEY_FLAG'] = 'M'
        
        tb = load_workbook(filename = 'template.xlsx')
        ts = tb.active
        self.fieldTemp={}
        for col in ts.iter_cols():
            self.fieldTemp[col[0].value] = col[0].column
            
    def start(self, startingRow = 2, limit = 0):
        print('Starting transformation')
        rowNum = startingRow
        temp = limit
        if temp >0:
            self.upperRange = limit
        while rowNum < self.upperRange:
            self.processRow(rowNum)
            rowNum+=1
        print('Complete! Saving File...')
        self.output.save(self.outputFileName)
        print('Done')
    def processRow(self,rowNum):
        print('Current Row is: '+str(rowNum))
        
        #check for change in info area
        data = self.ws[self.fieldSrc['Child']+str(rowNum)].value
        '''
        print(data)
        if data is None:
            print('lmaooooooooooooooooooooo')
        '''
        #data = str(data)
        #'''
        #print('Info area='+data)
        if (data == '') or (data is None): #initial state or no info area
            self.hasInfoArea = False
        elif (data != self.currentInfoArea) and (self.hasInfoArea == True): #not equal to current info area, change detected
            print('Creating new info area {info} with parent {par}'.format(info=data,par=self.ws[self.fieldSrc['Parent']+str(rowNum)].value))
            self.currentInfoArea = data
            self.createNewInfoArea(rowNum)
            self.createNewInfoArea(rowNum,isChild=True)
            self.hasInfoArea = True 
        else:
            self.hasInfoArea = False
        '''
            if (data is not None) and (data != ''): #and (self.ws[self.fieldSrc['Parent']+str(rowNum)].value is not None):
                print('Creating new info area {info} with parent {par}'.format(info=data,par=self.ws[self.fieldSrc['Parent']+str(rowNum)].value))
                self.currentInfoArea = data
                self.createNewInfoArea(rowNum)
                self.createNewInfoArea(rowNum,isChild=True)
                self.hasInfoArea = True 
            else:
                self.hasInfoArea = False
        '''
            
        #check for change in table
        data = self.ws[self.fieldSrc['DD_TABLENAME']+str(rowNum)].value
        #data = str(data)
        #print('TABLE = '+data)
        if (data == '') or (data is None): #initial state or no table
            print('No more of this table exists')
            self.hasTable = False
        elif data !=self.currentTable and len(data)>0: #mismatch detected
            self.hasTable = True
            self.currentTable = data
            self.createNewTable(rowNum)
        
        #build columns
        if data != '' and (data is not None) and self.hasTable: 
            if self.ws[self.fieldSrc['DD_FIELDNAME']+str(rowNum)].value is not None: #columns must be put in a table
                print('Creating column {col} under table: {tab}'.format(col='data',tab=self.ws[self.fieldSrc['DD_TABLENAME']+str(rowNum)].value))
                self.createNewColumn(rowNum)
        #'''
    
    def createNewInfoArea(self,rowNum, isChild = False):
        self.sOutput[self.fieldTemp['Status']+str(self.outputRowNum)] = 'Candidate'
        self.sOutput[self.fieldTemp['Type']+str(self.outputRowNum)] = 'Info Area'
        self.sOutput[self.fieldTemp['Domain']+str(self.outputRowNum)] = self.domain
        self.sOutput[self.fieldTemp['Community']+str(self.outputRowNum)] = self.community
        self.sOutput[self.fieldTemp['Domain Type']+str(self.outputRowNum)] = self.domainType

        if isChild:
            self.sOutput[self.fieldTemp['Name']+str(self.outputRowNum)] = self.ws[self.fieldSrc['Parent']+str(rowNum)].value + '::' + self.ws[self.fieldSrc['Child']+str(rowNum)].value
        
            #relation to parent
            self.sOutput[self.fieldTemp['is a child of [Info Area] > Info Area']+str(self.outputRowNum)] = self.ws[self.fieldSrc['Parent']+str(rowNum)].value
            self.sOutput[self.fieldTemp['is a child of [Info Area] > Type']+str(self.outputRowNum)] = 'Info Area'
            self.sOutput[self.fieldTemp['is a child of [Info Area] > Community']+str(self.outputRowNum)] = self.community
            self.sOutput[self.fieldTemp['is a child of [Info Area] > Domain Type']+str(self.outputRowNum)] = self.domainType
            self.sOutput[self.fieldTemp['is a child of [Info Area] > Domain']+str(self.outputRowNum)] = self.domain
        else:
            self.sOutput[self.fieldTemp['Name']+str(self.outputRowNum)] = self.ws[self.fieldSrc['Parent']+str(rowNum)].value
            
        self.outputRowNum +=1
           
    def createNewTable(self,rowNum):
        self.sOutput[self.fieldTemp['Status']+str(self.outputRowNum)] = 'Candidate'
        self.sOutput[self.fieldTemp['Type']+str(self.outputRowNum)] = 'Table'
        self.sOutput[self.fieldTemp['Domain']+str(self.outputRowNum)] = self.domain
        self.sOutput[self.fieldTemp['Community']+str(self.outputRowNum)] = self.community
        self.sOutput[self.fieldTemp['Domain Type']+str(self.outputRowNum)] = self.domainType
        
        self.sOutput[self.fieldTemp['Table Type']+str(self.outputRowNum)] = self.ws[self.fieldSrc['DD_TABLETYPE']+str(rowNum)].value
        self.sOutput[self.fieldTemp['Description']+str(self.outputRowNum)] = self.ws[self.fieldSrc['LONG_DESC']+str(rowNum)].value
        
        #relation (sometimes has no info area)
        if self.hasInfoArea and self.currentInfoArea !='' and self.currentInfoArea is not None:
            self.sOutput[self.fieldTemp['Name']+str(self.outputRowNum)] = self.ws[self.fieldSrc['Child']+str(rowNum)].value + '::' + self.ws[self.fieldSrc['DD_TABLENAME']+str(rowNum)].value
            self.sOutput[self.fieldTemp['is captured in [Info Area] > Info Area']+str(self.outputRowNum)] = self.ws[self.fieldSrc['Child']+str(rowNum)].value
            self.sOutput[self.fieldTemp['is captured in [Info Area] > Type']+str(self.outputRowNum)] = 'Info Area'
            self.sOutput[self.fieldTemp['is captured in [Info Area] > Community']+str(self.outputRowNum)] = self.community
            self.sOutput[self.fieldTemp['is captured in [Info Area] > Domain Type']+str(self.outputRowNum)] = self.domainType
            self.sOutput[self.fieldTemp['is captured in [Info Area] > Domain']+str(self.outputRowNum)] = self.domain
        else:
            self.sOutput[self.fieldTemp['Name']+str(self.outputRowNum)] = self.ws[self.fieldSrc['DD_TABLENAME']+str(rowNum)].value
        
        self.outputRowNum +=1
     
    def createNewColumn(self,rowNum):
        if self.hasInfoArea and self.currentInfoArea !='' and self.currentInfoArea is not None:
            self.sOutput[self.fieldTemp['Name']+str(self.outputRowNum)] =  self.ws[self.fieldSrc['Child']+str(rowNum)].value + '::' + self.ws[self.fieldSrc['DD_TABLENAME']+str(rowNum)].value + '::' + self.ws[self.fieldSrc['DD_FIELDNAME']+str(rowNum)].value
        else:
            self.sOutput[self.fieldTemp['Name']+str(self.outputRowNum)] = self.ws[self.fieldSrc['DD_TABLENAME']+str(rowNum)].value + '::' + self.ws[self.fieldSrc['DD_FIELDNAME']+str(rowNum)].value
        self.sOutput[self.fieldTemp['Status']+str(self.outputRowNum)] = 'Candidate'
        self.sOutput[self.fieldTemp['Type']+str(self.outputRowNum)] = 'Column'
        self.sOutput[self.fieldTemp['Domain']+str(self.outputRowNum)] = self.domain
        self.sOutput[self.fieldTemp['Community']+str(self.outputRowNum)] = self.community
        self.sOutput[self.fieldTemp['Domain Type']+str(self.outputRowNum)] = self.domainType
        
        #attributes
        self.sOutput[self.fieldTemp['Is Nullable']+str(self.outputRowNum)] = self.convertToCommonTerm(self.ws[self.fieldSrc['MANDATORY']+str(rowNum)].value)
        self.sOutput[self.fieldTemp['Description']+str(self.outputRowNum)] = self.ws[self.fieldSrc['SHORT_DESC']+str(rowNum)].value
        self.sOutput[self.fieldTemp['Is Primary Key']+str(self.outputRowNum)] = self.convertToCommonTerm(self.ws[self.fieldSrc['KEY_FLAG']+str(rowNum)].value)
        self.sOutput[self.fieldTemp['Number of Fractional Digits']+str(self.outputRowNum)] = self.ws[self.fieldSrc['DATA_DECIMALS']+str(rowNum)].value
        self.sOutput[self.fieldTemp['Size']+str(self.outputRowNum)] = self.ws[self.fieldSrc['DATA_LENGTH']+str(rowNum)].value
        self.sOutput[self.fieldTemp['Column Position']+str(self.outputRowNum)] = self.ws[self.fieldSrc['POSIT']+str(rowNum)].value
        self.sOutput[self.fieldTemp['Technical Data Type']+str(self.outputRowNum)] = self.ws[self.fieldSrc['DD_DATATYPE_ERP']+str(rowNum)].value
        
        #relation
        self.sOutput[self.fieldTemp['is part of [Table] > Table']+str(self.outputRowNum)] = self.ws[self.fieldSrc['DD_TABLENAME']+str(rowNum)].value
        self.sOutput[self.fieldTemp['is part of [Table] > Type']+str(self.outputRowNum)] = 'Table'
        self.sOutput[self.fieldTemp['is part of [Table] > Community']+str(self.outputRowNum)] = self.community
        self.sOutput[self.fieldTemp['is part of [Table] > Domain Type']+str(self.outputRowNum)] = self.domainType
        self.sOutput[self.fieldTemp['is part of [Table] > Domain']+str(self.outputRowNum)] = self.domain
            
        self.outputRowNum +=1
        
    def convertToCommonTerm(self,v):
        v = str(v)
        returnValue = v
        for c in ('yes', 'true', 't', 'y'):
            if v.lower() == c:
                returnValue = 'True'
        for c in ('no', 'false', 'f', 'n'):
            if v.lower() == c:
                returnValue = 'False'
        for c in ("none",'none'): # there is a difference between ' and " !
            if v.lower() == c:
                returnValue = 'False'
        for c in ("",''):
            if v.lower() == c:
                returnValue = 'False'
        return returnValue