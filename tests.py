import unittest
from openpyxl import load_workbook
import sys, os
sys.path.append(os.getcwd())
from transform import *

#crappiest unit tests you've probably ever seen #notevenadev
class TestTransform(unittest.TestCase):
    def setUp(self):
        self.t = SapDataParser('data.xlsx','output.xlsx')
    '''
    def testTryMappingColumns(self):
        wb = load_workbook('template.xlsx')
        ws = wb.active
        for col in ws.iter_cols():
            print(col[0].value)
    '''
    
    def testTransformData(self):
        self.t.start(limit = 49647)#limit = 49647     755
if __name__ == '__main__':
    unittest.main()